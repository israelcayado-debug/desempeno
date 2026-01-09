import csv
import io
import logging
import time
import re
from collections import Counter, defaultdict
from datetime import datetime

from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.http import HttpResponse, HttpResponseNotAllowed
from django.shortcuts import render, redirect
from django.db.models import Count, Max
from django.db import models
from django.urls import reverse
from django.http import QueryDict
from django.utils import timezone
from django.utils.text import Truncator

from apps.core.permissions import can_evaluate, is_hr_admin, is_manager
from apps.org.models import Position
from apps.org.selectors import employees_visible_to
from apps.evaluations.models import (
    EvaluationPeriod,
    Evaluation,
    EvaluationScore,
    EvaluationItem,
    ReportFilterPreset,
)
from apps.templates_eval.models import (
    EvaluationTemplate,
    TemplateActive,
    TemplateAssignment,
    TemplateQuestion,
)
from apps.templates_eval.services import resolve_active_template

try:
    import openpyxl
except Exception:  # pragma: no cover
    openpyxl = None

logger = logging.getLogger(__name__)


def compute_final_score(items):
    scores = []
    for item in items:
        if item.question_type == TemplateQuestion.SCALE_1_5 and item.value_scale is not None:
            scores.append(item.value_scale)
    if not scores:
        return None
    return round(sum(scores) / len(scores), 2)


def can_edit_evaluation(user, evaluation: Evaluation) -> bool:
    if evaluation.status == Evaluation.Status.DRAFT:
        return is_manager(user)
    if evaluation.status == Evaluation.Status.SUBMITTED:
        return False
    if evaluation.status == Evaluation.Status.FINAL:
        return False
    return False


def can_override_period_lock(user) -> bool:
    return is_hr_admin(user)


def can_view_reports(user) -> bool:
    return is_manager(user) or is_hr_admin(user)


def resolve_default_period():
    try:
        EvaluationPeriod._meta.get_field("is_active")
        active = EvaluationPeriod.objects.filter(is_active=True).first()
        if active:
            return active
    except Exception:
        pass

    open_qs = EvaluationPeriod.objects.filter(is_closed=False).order_by(
        "-end_date", "-start_date", "-id"
    )
    if open_qs.exists():
        return open_qs.first()

    return EvaluationPeriod.objects.order_by("-end_date", "-start_date", "-id").first()


def item_answer_as_text(item) -> str:
    if item.question_type == TemplateQuestion.SCALE_1_5:
        return "" if item.value_scale is None else str(item.value_scale)
    if item.question_type == TemplateQuestion.YES_NO:
        if item.value_yes_no is None:
            return ""
        return "YES" if item.value_yes_no else "NO"
    if item.question_type == TemplateQuestion.TEXT:
        return item.value_text or ""
    return ""


BLOCK_RE = re.compile(r"bloque\\s+([a-e])\\b", re.IGNORECASE)


def block_from_section(title: str) -> str:
    m = BLOCK_RE.search(title or "")
    return m.group(1).upper() if m else "UNK"


def create_items_from_template(evaluation: Evaluation, template) -> None:
    items = []
    order = 1
    sections = template.sections.prefetch_related("questions").order_by("order", "id")
    for section in sections:
        questions = section.questions.order_by("order", "id")
        for q in questions:
            is_required = getattr(q, "is_required", None)
            if is_required is None:
                is_required = q.required
            items.append(
                EvaluationItem(
                    evaluation=evaluation,
                    section_title=section.title,
                    question_text=q.text,
                    question_type=q.question_type,
                    is_required=bool(is_required),
                    display_order=order,
                )
            )
            order += 1
    if items:
        EvaluationItem.objects.bulk_create(items)


def compute_block_scores(items):
    by_block = defaultdict(list)
    for item in items:
        if item.question_type != TemplateQuestion.SCALE_1_5:
            continue
        if item.value_scale is None:
            continue
        by_block[block_from_section(item.section_title)].append(item.value_scale)

    scores = {}
    for block, values in by_block.items():
        scores[block] = round(sum(values) / len(values), 2) if values else None
    return scores


def build_period_report_queryset(request, period, user):
    employees = employees_visible_to(user)
    qs = (
        Evaluation.objects.filter(period=period, employee__in=employees)
        .select_related("employee", "period")
    )
    status = (request.GET.get("status") or "").strip().upper()
    if status in {Evaluation.Status.DRAFT, Evaluation.Status.SUBMITTED, Evaluation.Status.FINAL}:
        qs = qs.filter(status=status)

    q = (request.GET.get("q") or "").strip()
    if q:
        qs = qs.filter(
            models.Q(employee__full_name__icontains=q)
            | models.Q(employee__dni__icontains=q)
        )

    sort_map = {
        "employee": "employee__full_name",
        "status": "status",
        "score": "final_score",
        "updated": "status_changed_at",
        "submitted": "submitted_at",
        "finalized": "finalized_at",
    }
    sort = (request.GET.get("sort") or "employee").strip().lower()
    direction = (request.GET.get("dir") or "asc").strip().lower()
    sort_field = sort_map.get(sort, "employee__full_name")
    if direction not in {"asc", "desc"}:
        direction = "asc"
    if direction == "desc":
        sort_field = f"-{sort_field}"

    qs = qs.order_by(sort_field, "employee__id")
    return qs, {"status": status, "q": q, "sort": sort, "dir": direction}


def get_block_codes(items):
    seen = set()
    ordered = []
    for item in items:
        code = block_from_section(item.section_title)
        if code not in seen:
            ordered.append(code)
            seen.add(code)
    priority = ["A", "B", "C", "D", "E"]
    ordered_sorted = [c for c in priority if c in seen]
    for code in ordered:
        if code not in ordered_sorted:
            ordered_sorted.append(code)
    return ordered_sorted


def build_querystring(request, *, exclude=None, overrides=None) -> str:
    exclude = set(exclude or [])
    qd = QueryDict("", mutable=True)
    for key, values in request.GET.lists():
        if key in exclude:
            continue
        for v in values:
            qd.appendlist(key, v)
    if overrides:
        for key, value in overrides.items():
            if value is None:
                qd.pop(key, None)
            else:
                qd[key] = str(value)
    return qd.urlencode()


def normalize_filters(request):
    items = []
    for key in sorted(request.GET.keys()):
        values = request.GET.getlist(key)
        if len(values) == 1:
            items.append((key, values[0]))
        else:
            items.append((key, values))
    return items


def user_role_label(user) -> str:
    if is_hr_admin(user):
        return "HR_ADMIN"
    if is_manager(user):
        return "MANAGER"
    return "OTHER"



@login_required
def my_team(request):
    if not can_evaluate(request.user):
        raise PermissionDenied

    employees = (
        employees_visible_to(request.user)
        .select_related("evaluation_position", "manager")
        .order_by("full_name")
    )

    current_period = (
        EvaluationPeriod.objects.filter(is_closed=False)
        .order_by("-start_date")
        .first()
    )
    eval_by_employee_id = {}
    if current_period:
        evals = (
            Evaluation.objects
            .filter(period=current_period, employee__in=employees)
            .only("id", "employee_id", "status", "final_score", "overall_comment")
        )
        eval_by_employee_id = {ev.employee_id: ev for ev in evals}

    alerts_by_employee_id = {}
    today = timezone.now().date()
    if current_period:
        for e in employees:
            ev = eval_by_employee_id.get(e.id)
            alerts = []
            if not ev:
                alerts.append(
                    {
                        "code": "NOT_STARTED",
                        "severity": "info",
                        "text": "Sin iniciar",
                    }
                )
            else:
                is_sent = ev.status in {Evaluation.Status.SUBMITTED, Evaluation.Status.FINAL}
                is_overdue = current_period.end_date and today > current_period.end_date and not is_sent
                if is_overdue:
                    alerts.append(
                        {
                            "code": "OVERDUE",
                            "severity": "danger",
                            "text": "Vencida",
                        }
                    )
                if ev.status == Evaluation.Status.DRAFT and not is_overdue:
                    alerts.append(
                        {
                            "code": "DRAFT",
                            "severity": "warning",
                            "text": "Borrador",
                        }
                    )
                if not (ev.overall_comment or "").strip():
                    alerts.append(
                        {
                            "code": "MISSING_REQUIRED",
                            "severity": "warning",
                            "text": "Faltan obligatorios",
                        }
                    )
                if not can_edit_evaluation(request.user, ev):
                    alerts.append(
                        {
                            "code": "BLOCKED",
                            "severity": "info",
                            "text": "Bloqueada",
                        }
                    )

            for a in alerts:
                a["action_url"] = (
                    f"/evaluate/{e.id}/{current_period.id}/" if current_period else ""
                )
            alerts_by_employee_id[e.id] = alerts

    position_ids = [e.evaluation_position_id for e in employees if e.evaluation_position_id]

    assignments = (
        TemplateAssignment.objects
        .select_related("template", "position")
        .filter(position_id__in=position_ids, template__is_active=True)
        .order_by("position_id", "-is_default", "-template__version")
    )

    template_by_position_id = {}
    for a in assignments:
        template_by_position_id.setdefault(a.position_id, a.template)

    return render(
        request,
        "evaluations/my_team.html",
        {
            "employees": employees,
            "current_period": current_period,
            "template_by_position_id": template_by_position_id,
            "eval_by_employee_id": eval_by_employee_id,
            "alerts_by_employee_id": alerts_by_employee_id,

        },
    )


@login_required
def evaluate_employee(request, employee_id: int, period_id: int):
    if not can_evaluate(request.user):
        raise PermissionDenied

    employee = (
        employees_visible_to(request.user)
        .filter(id=employee_id)
        .select_related("evaluation_position")
        .first()
    )
    if not employee:
        raise PermissionDenied

    period = EvaluationPeriod.objects.filter(id=period_id).first()
    if not period:
        raise PermissionDenied

    override = request.GET.get("override") == "1"
    if request.method == "POST" and request.POST.get("override") == "1":
        override = True
    period_locked = bool(period.is_closed)
    override_allowed = can_override_period_lock(request.user)

    if not employee.evaluation_position:
        return render(
            request,
            "evaluations/evaluate_employee.html",
            {
                "employee": employee,
                "period": period,
                        "period_locked": period_locked,
                        "override_allowed": override_allowed,
                        "override": override,
                "error": "El empleado no tiene Puesto de EvaluaciÃ³n asignado.",
            },
        )

    base_code = employee.evaluation_position.code

    tpl = resolve_active_template(base_code)
    if not tpl:
        return render(
            request,
            "evaluations/evaluate_employee.html",
            {
                "employee": employee,
                "period": period,
                        "period_locked": period_locked,
                        "override_allowed": override_allowed,
                        "override": override,
                "error": f"No hay plantilla activa para {base_code}.",
            },
        )

    evaluation = Evaluation.objects.filter(employee=employee, period=period).first()
    created = False
    if evaluation is None:
        if period_locked and not (override and override_allowed):
            return render(
                request,
                "evaluations/evaluate_employee.html",
                {
                    "employee": employee,
                    "period": period,
                    "period_locked": period_locked,
                    "override_allowed": override_allowed,
                    "override": override,
                    "error": "El periodo esta cerrado. No se pueden crear nuevas evaluaciones.",
                },
            )
        evaluation = Evaluation.objects.create(
            employee=employee,
            period=period,
            evaluator=request.user,
            template=tpl,
            frozen_position_code=base_code,
            frozen_position_name=employee.evaluation_position.name,
        )
        created = True

    if evaluation.template_id is None:
        tpl = resolve_active_template(base_code)
        if not tpl:
            return render(
                request,
                "evaluations/evaluate_employee.html",
                {
                    "employee": employee,
                    "period": period,
                        "period_locked": period_locked,
                        "override_allowed": override_allowed,
                        "override": override,
                    "error": f"No hay plantilla activa para {base_code}.",
                },
            )
        evaluation.template = tpl
        evaluation.save(update_fields=["template"])

    template = evaluation.template
    if created and template:
        create_items_from_template(evaluation, template)
    items = list(evaluation.items.all().order_by("display_order", "id"))
    block_codes = get_block_codes(items)
    blocks = [{"code": code} for code in block_codes]

    history_qs = (
        Evaluation.objects.filter(employee=employee)
        .exclude(id=evaluation.id)
        .select_related("period")
        .order_by("-period__start_date", "-updated_at")
    )[:12]
    history = []
    for ev in history_qs:
        history.append(
            {
                "id": ev.id,
                "period_name": ev.period.name,
                "period_start": ev.period.start_date,
                "period_end": ev.period.end_date,
                "status": ev.get_status_display(),
                "updated_at": ev.updated_at,
                "overall_comment": Truncator((ev.overall_comment or "").strip()).chars(120),
            }
        )

    error = None
    editable = can_edit_evaluation(request.user, evaluation)
    can_finalize = is_hr_admin(request.user)
    if period_locked and not (override and override_allowed):
        editable = False

    def is_item_complete(item) -> bool:
        qtype = item.question_type
        if qtype == TemplateQuestion.SCALE_1_5:
            return item.value_scale is not None
        if qtype == TemplateQuestion.YES_NO:
            return item.value_yes_no is not None
        if qtype == TemplateQuestion.TEXT:
            return bool((item.value_text or "").strip())
        return False

    def get_missing_required_item_ids(items) -> set:
        missing_required = set()
        for item in items:
            if not item.is_required:
                continue
            if not is_item_complete(item):
                missing_required.add(item.id)
        return missing_required

    if request.method == "POST":
        action = request.POST.get("action", "save")

        if period_locked and not (override and override_allowed):
            return render(
                request,
                "evaluations/evaluate_employee.html",
                {
                    "employee": employee,
                    "period": period,
                    "period_locked": period_locked,
                    "override_allowed": override_allowed,
                    "override": override,
                    "evaluation": evaluation,
                    "created": created,
                    "template": template,
                    "items": items,
                    "missing_required": set(),
                    "can_close": can_finalize,
                    "editable": editable,
                    "error": "El periodo esta cerrado. Acciones deshabilitadas (usa override si procede).",
                },
            )

        if action in ("save", None, ""):
            if not editable:
                return render(
                    request,
                    "evaluations/evaluate_employee.html",
                    {
                        "employee": employee,
                        "period": period,
                        "period_locked": period_locked,
                        "override_allowed": override_allowed,
                        "override": override,
                        "evaluation": evaluation,
                        "created": created,
                        "template": template,
                        "items": items,
                        "blocks": blocks,
                        "missing_required": set(),
                        "can_close": can_finalize,
                        "editable": editable,
                        "error": "Esta evaluacion no es editable en su estado actual.",
                    },
                )

        if action == "finalize":
            if not can_finalize:
                return render(
                    request,
                    "evaluations/evaluate_employee.html",
                    {
                        "employee": employee,
                        "period": period,
                        "period_locked": period_locked,
                        "override_allowed": override_allowed,
                        "override": override,
                        "evaluation": evaluation,
                        "created": created,
                        "template": template,
                        "items": items,
                        "blocks": blocks,
                        "missing_required": set(),
                        "can_close": can_finalize,
                        "editable": editable,
                        "error": "No tienes permisos para cerrar la evaluacion.",
                    },
                )
            if evaluation.status != Evaluation.Status.SUBMITTED:
                return render(
                    request,
                    "evaluations/evaluate_employee.html",
                    {
                        "employee": employee,
                        "period": period,
                        "period_locked": period_locked,
                        "override_allowed": override_allowed,
                        "override": override,
                        "evaluation": evaluation,
                        "created": created,
                        "template": template,
                        "items": items,
                        "blocks": blocks,
                        "missing_required": set(),
                        "can_close": can_finalize,
                        "editable": editable,
                        "error": "Solo se puede cerrar una evaluacion enviada.",
                    },
                )
            evaluation.set_status(Evaluation.Status.FINAL)
            evaluation.save(
                update_fields=["status", "finalized_at", "status_changed_at"]
            )
            return redirect("evaluate_employee", employee_id=employee.id, period_id=period.id)

        if action == "reopen":
            if not can_finalize:
                return render(
                    request,
                    "evaluations/evaluate_employee.html",
                    {
                        "employee": employee,
                        "period": period,
                        "period_locked": period_locked,
                        "override_allowed": override_allowed,
                        "override": override,
                        "evaluation": evaluation,
                        "created": created,
                        "template": template,
                        "items": items,
                        "blocks": blocks,
                        "missing_required": set(),
                        "can_close": can_finalize,
                        "editable": editable,
                        "error": "No tienes permisos para reabrir una evaluacion cerrada.",
                    },
                )
            if evaluation.status != Evaluation.Status.FINAL:
                return render(
                    request,
                    "evaluations/evaluate_employee.html",
                    {
                        "employee": employee,
                        "period": period,
                        "period_locked": period_locked,
                        "override_allowed": override_allowed,
                        "override": override,
                        "evaluation": evaluation,
                        "created": created,
                        "template": template,
                        "items": items,
                        "blocks": blocks,
                        "missing_required": set(),
                        "can_close": can_finalize,
                        "editable": editable,
                        "error": "Solo se pueden reabrir evaluaciones en estado FINAL.",
                    },
                )
            reason = (request.POST.get("reopen_reason") or "").strip()
            if not reason:
                return render(
                    request,
                    "evaluations/evaluate_employee.html",
                    {
                        "employee": employee,
                        "period": period,
                        "period_locked": period_locked,
                        "override_allowed": override_allowed,
                        "override": override,
                        "evaluation": evaluation,
                        "created": created,
                        "template": template,
                        "items": items,
                        "blocks": blocks,
                        "missing_required": set(),
                        "can_close": can_finalize,
                        "editable": editable,
                        "error": "Para reabrir una evaluacion es obligatorio indicar un motivo.",
                    },
                )
            evaluation.set_status(Evaluation.Status.DRAFT, reason=reason)
            evaluation.save(
                update_fields=["status", "reopened_at", "reopen_reason", "status_changed_at"]
            )
            return redirect("evaluate_employee", employee_id=employee.id, period_id=period.id)

        previous_status = evaluation.status

        if action in ("save", "submit") and editable:
            comment = request.POST.get("evaluator_comment", "").strip()
            if evaluation.evaluator_comment != comment:
                evaluation.evaluator_comment = comment
            overall_comment = (request.POST.get("overall_comment") or "").strip()
            if evaluation.overall_comment != overall_comment:
                evaluation.overall_comment = overall_comment

            # 1) Guardar respuestas
            for item in items:
                key = f"q_{item.id}"
                if key not in request.POST:
                    continue

                val = request.POST.get(key)

                if item.question_type == TemplateQuestion.SCALE_1_5:
                    item.value_scale = int(val) if val else None
                    item.value_yes_no = None
                    item.value_text = None

                elif item.question_type == TemplateQuestion.YES_NO:
                    item.value_yes_no = True if val == "1" else False
                    item.value_scale = None
                    item.value_text = None

                elif item.question_type == TemplateQuestion.TEXT:
                    item.value_text = val or ""
                    item.value_scale = None
                    item.value_yes_no = None

                item.save()

        # 2) Si action=submit, cambiar estado
        final_score_set = False
        if action == "submit":
            if not editable:
                return render(
                    request,
                    "evaluations/evaluate_employee.html",
                    {
                        "employee": employee,
                        "period": period,
                        "period_locked": period_locked,
                        "override_allowed": override_allowed,
                        "override": override,
                        "evaluation": evaluation,
                        "created": created,
                        "template": template,
                        "items": items,
                        "blocks": blocks,
                        "missing_required": set(),
                        "can_close": can_finalize,
                        "editable": editable,
                        "error": "No tienes permisos para enviar esta evaluacion.",
                    },
                )
            if evaluation.status != Evaluation.Status.DRAFT:
                return redirect("evaluate_employee", employee_id=employee.id, period_id=period.id)
            missing = []
            for item in items:
                if not item.is_required:
                    continue
                if not is_item_complete(item):
                    missing.append((item.section_title, item.question_text))

            if missing:
                lines = [f"- {sec}: {q}" for sec, q in missing[:15]]
                suffix = "" if len(missing) <= 15 else f"\n(...y {len(missing) - 15} mas)"
                error = (
                    "No se puede enviar. Faltan respuestas obligatorias:\n"
                    + "\n".join(lines)
                    + suffix
                )
            else:
                final = compute_final_score(items)
                evaluation.final_score = final
                final_score_set = True
                evaluation.set_status(Evaluation.Status.SUBMITTED)

        update_fields = []
        if action in ("save", "submit") and editable:
            update_fields.append("evaluator_comment")
            update_fields.append("overall_comment")
        if evaluation.status != previous_status:
            update_fields.extend(
                ["status", "submitted_at", "finalized_at", "reopened_at", "reopen_reason", "status_changed_at"]
            )
        if final_score_set:
            update_fields.append("final_score")
        if update_fields:
            evaluation.save(update_fields=update_fields)

        if error is None:
            return redirect("evaluate_employee", employee_id=employee.id, period_id=period.id)

    items = list(evaluation.items.all().order_by("display_order", "id"))
    block_codes = get_block_codes(items)
    blocks = [{"code": code} for code in block_codes]
    if evaluation.status == Evaluation.Status.DRAFT and editable:
        missing_required = get_missing_required_item_ids(items)
    else:
        missing_required = set()
    score_total = compute_final_score(items)
    block_scores = compute_block_scores(items)
    pending_required_count = len(missing_required)
    today = timezone.now().date()
    is_overdue = (
        period.end_date is not None
        and today > period.end_date
        and evaluation.status not in {Evaluation.Status.SUBMITTED, Evaluation.Status.FINAL}
    )
    is_incomplete = pending_required_count > 0 or not (evaluation.overall_comment or "").strip()

    return render(
        request,
        "evaluations/evaluate_employee.html",
        {
            "employee": employee,
            "period": period,
                        "period_locked": period_locked,
                        "override_allowed": override_allowed,
                        "override": override,
            "evaluation": evaluation,
            "created": created,
            "template": template,
            "items": items,
            "blocks": blocks,
            "history": history,
            "missing_required": missing_required,
            "can_close": can_finalize,
            "editable": editable,
            "score_total": score_total,
            "block_scores": block_scores,
            "pending_required_count": pending_required_count,
            "is_overdue": is_overdue,
            "is_incomplete": is_incomplete,
            "is_history": False,
            "error": error,


        },
    )


@login_required
def evaluation_history_view(request, evaluation_id: int):
    if request.method != "GET":
        return HttpResponseNotAllowed(["GET"])
    if not can_evaluate(request.user):
        raise PermissionDenied

    ev = Evaluation.objects.select_related("employee", "period").filter(id=evaluation_id).first()
    if not ev:
        raise PermissionDenied

    visible = employees_visible_to(request.user).filter(id=ev.employee_id).exists()
    if not visible:
        raise PermissionDenied

    items = list(ev.items.all().order_by("display_order", "id"))
    block_codes = get_block_codes(items)
    blocks = [{"code": code} for code in block_codes]
    score_total = compute_final_score(items)
    block_scores = compute_block_scores(items)
    pending_required_count = 0

    return render(
        request,
        "evaluations/evaluate_employee.html",
        {
            "employee": ev.employee,
            "period": ev.period,
            "evaluation": ev,
            "created": False,
            "template": ev.template,
            "items": items,
            "blocks": blocks,
            "history": [],
            "missing_required": set(),
            "can_close": False,
            "editable": False,
            "score_total": score_total,
            "block_scores": block_scores,
            "pending_required_count": pending_required_count,
            "is_history": True,
            "period_locked": False,
            "override_allowed": False,
            "override": False,
            "error": None,
        },
    )


@login_required
def report_period(request, period_id: int | None = None):
    if not can_view_reports(request.user):
        raise PermissionDenied

    if request.method == "POST" and is_hr_admin(request.user):
        action = request.POST.get("action")
        if action == "save_preset":
            name = (request.POST.get("preset_name") or "").strip()
            if name:
                preset_period = period_id or request.GET.get("period")
                overrides = {}
                if preset_period:
                    overrides["period"] = str(preset_period)
                query = build_querystring(
                    request,
                    exclude={"page", "export_scope"},
                    overrides=overrides,
                )
                ReportFilterPreset.objects.create(
                    name=name,
                    scope="period_dashboard",
                    query_params=query,
                    created_by=request.user,
                )
        elif action == "delete_preset":
            preset_id = request.POST.get("preset_id")
            preset = ReportFilterPreset.objects.filter(
                id=preset_id, scope="period_dashboard"
            ).first()
            if preset and (preset.created_by_id == request.user.id or is_hr_admin(request.user)):
                preset.delete()
        return redirect(request.path)

    periods = list(EvaluationPeriod.objects.order_by("-end_date", "-start_date", "-id"))
    show_period_selector = True
    period = None
    if period_id is not None:
        period = EvaluationPeriod.objects.filter(id=period_id).first()
    else:
        if request.GET.get("period"):
            period = EvaluationPeriod.objects.filter(id=request.GET.get("period")).first()
        if not period:
            period = resolve_default_period()

    if not period:
        return render(
            request,
            "evaluations/report_period.html",
            {
                "periods": periods,
                "period": None,
                "totals": {},
                "evaluations": [],
                "status_filter": "",
                "search": "",
                "show_period_selector": show_period_selector,
            },
        )

    employees = employees_visible_to(request.user)
    base_qs = (
        Evaluation.objects.filter(period=period, employee__in=employees)
        .select_related("employee", "period")
    )

    totals = {row["status"]: row["count"] for row in base_qs.values("status").annotate(count=Count("id"))}
    totals_default = {
        "DRAFT": totals.get(Evaluation.Status.DRAFT, 0),
        "SUBMITTED": totals.get(Evaluation.Status.SUBMITTED, 0),
        "FINAL": totals.get(Evaluation.Status.FINAL, 0),
        "TOTAL": base_qs.count(),
    }

    qs, applied_filters = build_period_report_queryset(request, period, request.user)
    status_filter = applied_filters["status"]
    search = applied_filters["q"]
    sort = applied_filters["sort"]
    direction = applied_filters["dir"]

    page_size = request.GET.get("page_size") or "25"
    if page_size not in {"25", "50", "100"}:
        page_size = "25"
    page_number = request.GET.get("page") or "1"
    paginator = Paginator(qs, int(page_size))
    page_obj = paginator.get_page(page_number)
    evaluations = list(page_obj.object_list)
    filtered_count = paginator.count

    base_qs = build_querystring(request, exclude={"page"})
    sort_qs = build_querystring(request, exclude={"page", "sort", "dir"})
    export_qs_filtered = build_querystring(
        request,
        exclude={"page", "export_scope"},
        overrides={"export_scope": "filtered"},
    )
    export_qs_page = build_querystring(
        request,
        exclude={"export_scope"},
        overrides={"export_scope": "page"},
    )

    return render(
        request,
        "evaluations/report_period.html",
        {
            "periods": periods,
            "period": period,
            "totals": totals_default,
            "evaluations": evaluations,
            "status_filter": status_filter,
            "search": search,
            "filtered_count": filtered_count,
            "show_period_selector": show_period_selector,
            "page_obj": page_obj,
            "page_size": int(page_size),
            "sort": sort,
            "dir": direction,
            "base_qs": base_qs,
            "sort_qs": sort_qs,
            "export_qs_filtered": export_qs_filtered,
            "export_qs_page": export_qs_page,
            "presets": ReportFilterPreset.objects.filter(
                scope="period_dashboard"
            ).filter(
                models.Q(created_by=request.user) | models.Q(is_shared=True)
            ).order_by("name")
            if is_hr_admin(request.user)
            else [],
        },
    )


@login_required
def report_period_export_csv(request, period_id: int):
    if not can_view_reports(request.user):
        raise PermissionDenied

    period = EvaluationPeriod.objects.filter(id=period_id).first()
    if not period:
        raise PermissionDenied

    qs, _ = build_period_report_queryset(request, period, request.user)
    export_scope = (request.GET.get("export_scope") or "filtered").strip().lower()
    if export_scope == "page":
        page_size = request.GET.get("page_size") or "25"
        if page_size not in {"25", "50", "100"}:
            page_size = "25"
        page_number = request.GET.get("page") or "1"
        paginator = Paginator(qs, int(page_size))
        page_obj = paginator.get_page(page_number)
        qs = page_obj.object_list

    start = time.monotonic()
    output = io.StringIO()
    output.write("\ufeff")
    writer = csv.writer(output)
    writer.writerow(
        [
            "period_name",
            "employee_dni",
            "employee_full_name",
            "position_code",
            "status",
            "score_total",
            "overall_comment",
            "submitted_at",
            "finalized_at",
            "reopened_at",
        ]
    )
    evals = list(qs)

    row_count = 0
    for ev in evals:
        writer.writerow(
            [
                period.name,
                ev.employee.dni,
                ev.employee.full_name,
                ev.frozen_position_code,
                ev.status,
                ev.final_score or "",
                ev.overall_comment or "",
                ev.submitted_at or "",
                ev.finalized_at or "",
                ev.reopened_at or "",
            ]
        )
        row_count += 1

    resp = HttpResponse(output.getvalue(), content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="period_{period.id}_summary.csv"'
    logger.info(
        "report_export",
        extra={
            "event": "report_export",
            "export_type": "csv_summary",
            "period_id": period.id,
            "export_scope": export_scope,
            "rows": row_count,
            "user_id": request.user.id,
            "user_role": user_role_label(request.user),
            "filters": normalize_filters(request),
            "duration_ms": int((time.monotonic() - start) * 1000),
        },
    )
    return resp


@login_required
def report_period_export_items_csv(request, period_id: int):
    if not can_view_reports(request.user):
        raise PermissionDenied

    period = EvaluationPeriod.objects.filter(id=period_id).first()
    if not period:
        raise PermissionDenied

    qs, _ = build_period_report_queryset(request, period, request.user)
    export_scope = (request.GET.get("export_scope") or "filtered").strip().lower()
    if export_scope == "page":
        page_size = request.GET.get("page_size") or "25"
        if page_size not in {"25", "50", "100"}:
            page_size = "25"
        page_number = request.GET.get("page") or "1"
        paginator = Paginator(qs, int(page_size))
        page_obj = paginator.get_page(page_number)
        eval_ids = [ev.id for ev in page_obj.object_list]
    else:
        eval_ids = qs.values_list("id", flat=True)
    items_qs = EvaluationItem.objects.filter(evaluation_id__in=eval_ids).select_related(
        "evaluation", "evaluation__employee"
    )

    start = time.monotonic()
    output = io.StringIO()
    output.write("\ufeff")
    writer = csv.writer(output)
    writer.writerow(
        [
            "period",
            "employee_dni",
            "employee_full_name",
            "position_code",
            "evaluation_status",
            "section",
            "question_text",
            "question_type",
            "is_required",
            "answer_value",
            "display_order",
        ]
    )
    row_count = 0
    for item in items_qs.order_by("evaluation_id", "display_order", "id"):
        ev = item.evaluation
        writer.writerow(
            [
                period.name,
                ev.employee.dni,
                ev.employee.full_name,
                ev.frozen_position_code,
                ev.status,
                item.section_title,
                item.question_text,
                item.question_type,
                int(item.is_required),
                item_answer_as_text(item),
                item.display_order,
            ]
        )
        row_count += 1

    resp = HttpResponse(output.getvalue(), content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="period_{period.id}_items.csv"'
    logger.info(
        "report_export",
        extra={
            "event": "report_export",
            "export_type": "csv_items",
            "period_id": period.id,
            "export_scope": export_scope,
            "rows": row_count,
            "user_id": request.user.id,
            "user_role": user_role_label(request.user),
            "filters": normalize_filters(request),
            "duration_ms": int((time.monotonic() - start) * 1000),
        },
    )
    return resp


@login_required
def report_period_export_xlsx(request, period_id: int):
    if not can_view_reports(request.user):
        raise PermissionDenied
    if openpyxl is None:
        raise PermissionDenied

    period = EvaluationPeriod.objects.filter(id=period_id).first()
    if not period:
        raise PermissionDenied

    eval_qs, _ = build_period_report_queryset(request, period, request.user)
    export_scope = (request.GET.get("export_scope") or "filtered").strip().lower()
    if export_scope == "page":
        page_size = request.GET.get("page_size") or "25"
        if page_size not in {"25", "50", "100"}:
            page_size = "25"
        page_number = request.GET.get("page") or "1"
        paginator = Paginator(eval_qs, int(page_size))
        page_obj = paginator.get_page(page_number)
        eval_qs = page_obj.object_list
        eval_ids = [ev.id for ev in eval_qs]
    else:
        eval_ids = eval_qs.values_list("id", flat=True)
    start = time.monotonic()
    items_qs = EvaluationItem.objects.filter(evaluation_id__in=eval_ids).select_related(
        "evaluation", "evaluation__employee"
    )
    items_count = items_qs.count()
    if items_count > 50000:
        logger.info(
            "report_export",
            extra={
                "event": "report_export",
                "export_type": "xlsx_blocked",
                "period_id": period.id,
                "export_scope": export_scope,
                "items_count": items_count,
                "user_id": request.user.id,
                "user_role": user_role_label(request.user),
                "filters": normalize_filters(request),
                "duration_ms": int((time.monotonic() - start) * 1000),
            },
        )
        return HttpResponse(
            "Demasiados registros para XLSX. Usa CSV.",
            status=400,
            content_type="text/plain; charset=utf-8",
        )
    if items_count > 10000 and request.GET.get("confirm") != "1":
        link = f"{reverse('report_period_export_xlsx', args=[period.id])}?" + build_querystring(
            request,
            overrides={"confirm": "1"},
        )
        logger.info(
            "report_export",
            extra={
                "event": "report_export",
                "export_type": "xlsx_confirm",
                "period_id": period.id,
                "export_scope": export_scope,
                "items_count": items_count,
                "user_id": request.user.id,
                "user_role": user_role_label(request.user),
                "filters": normalize_filters(request),
                "duration_ms": int((time.monotonic() - start) * 1000),
            },
        )
        return HttpResponse(
            f'XLSX con {items_count} items. <a href="{link}">Continuar de todos modos</a>',
            content_type="text/html; charset=utf-8",
        )

    def to_naive(value):
        if value is None:
            return ""
        if timezone.is_aware(value):
            return timezone.make_naive(value)
        return value

    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Resumen"
    ws_summary.append(
        [
            "period_name",
            "employee_dni",
            "employee_full_name",
            "position_code",
            "status",
            "score_total",
            "overall_comment",
            "submitted_at",
            "finalized_at",
            "reopened_at",
        ]
    )
    evals = list(eval_qs)
    for ev in evals:
        ws_summary.append(
            [
                period.name,
                ev.employee.dni,
                ev.employee.full_name,
                ev.frozen_position_code,
                ev.status,
                ev.final_score or "",
                ev.overall_comment or "",
                to_naive(ev.submitted_at),
                to_naive(ev.finalized_at),
                to_naive(ev.reopened_at),
            ]
        )

    ws_detail = wb.create_sheet("Detalle")
    ws_detail.append(
        [
            "period",
            "employee_dni",
            "employee_full_name",
            "position_code",
            "evaluation_status",
            "section",
            "question_text",
            "question_type",
            "is_required",
            "answer_value",
            "display_order",
        ]
    )
    for item in items_qs.order_by("evaluation_id", "display_order", "id"):
        ev = item.evaluation
        ws_detail.append(
            [
                period.name,
                ev.employee.dni,
                ev.employee.full_name,
                ev.frozen_position_code,
                ev.status,
                item.section_title,
                item.question_text,
                item.question_type,
                int(item.is_required),
                item_answer_as_text(item),
                item.display_order,
            ]
        )

    ws_stats = wb.create_sheet("Stats")
    ws_stats.append(
        [
            "position_code",
            "block",
            "avg_score",
            "min_score",
            "max_score",
            "evaluations_count",
        ]
    )
    stats = defaultdict(list)
    evals_by_pos = defaultdict(set)
    for item in items_qs:
        if item.question_type != TemplateQuestion.SCALE_1_5 or item.value_scale is None:
            continue
        ev = item.evaluation
        block = block_from_section(item.section_title)
        key = (ev.frozen_position_code, block)
        stats[key].append(item.value_scale)
        evals_by_pos[key].add(ev.id)

    for (pos_code, block), values in sorted(stats.items()):
        ws_stats.append(
            [
                pos_code,
                block,
                round(sum(values) / len(values), 2) if values else "",
                min(values) if values else "",
                max(values) if values else "",
                len(evals_by_pos[(pos_code, block)]),
            ]
        )

    output = io.BytesIO()
    wb.save(output)
    resp = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet; charset=utf-8",
    )
    resp["Content-Disposition"] = f'attachment; filename="period_{period.id}.xlsx"'
    logger.info(
        "report_export",
        extra={
            "event": "report_export",
            "export_type": "xlsx",
            "period_id": period.id,
            "export_scope": export_scope,
            "items_count": items_count,
            "user_id": request.user.id,
            "user_role": user_role_label(request.user),
            "filters": normalize_filters(request),
            "duration_ms": int((time.monotonic() - start) * 1000),
        },
    )
    return resp


@login_required
def report_system(request):
    if not is_hr_admin(request.user):
        raise PermissionDenied

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    actives = TemplateActive.objects.select_related("template").all()
    active_templates = [a.template for a in actives if a.template_id]
    active_count = len(active_templates)
    active_base_codes = {t.base_code for t in active_templates if t.base_code}

    positions = list(Position.objects.all())
    position_count = len(positions)

    assignments = TemplateAssignment.objects.select_related("template").all()
    assign_map = {a.position_id: a.template.base_code for a in assignments}
    correct_assignments = sum(
        1 for p in positions if assign_map.get(p.id) in active_base_codes
    )

    latest_versions = (
        EvaluationTemplate.objects.values("base_code")
        .annotate(latest=Max("version"))
    )
    latest_map = {r["base_code"]: r["latest"] for r in latest_versions}

    outdated = []
    for t in active_templates:
        latest = latest_map.get(t.base_code)
        if latest is not None and latest != t.version:
            outdated.append((t.base_code, t.version, latest))

    q = TemplateQuestion.objects.filter(section__template__in=active_templates)
    q_types = sorted(set(q.values_list("question_type", flat=True)))
    q_counts = Counter(q.values_list("section__template__base_code", flat=True))
    count_values = sorted(set(q_counts.values()))
    count_min = min(count_values) if count_values else 0
    count_max = max(count_values) if count_values else 0

    missing_assignment = [p.code for p in positions if p.id not in assign_map]
    assigned_not_active = [
        (p.code, assign_map[p.id])
        for p in positions
        if p.id in assign_map and assign_map[p.id] not in active_base_codes
    ]

    position_codes = {p.code for p in positions}
    all_template_base_codes = set(
        EvaluationTemplate.objects.exclude(base_code="")
        .values_list("base_code", flat=True)
        .distinct()
    )
    legacy_base_codes = sorted(all_template_base_codes - position_codes)

    unexpected_types = [t for t in q_types if t != "SCALE_1_5"]
    issues = bool(outdated or missing_assignment or assigned_not_active or unexpected_types)

    if request.GET.get("csv") == "1":
        output = io.StringIO()
        output.write("\ufeff")
        writer = csv.writer(output)
        writer.writerow(["metric", "value"])
        writer.writerow(["active_templates", active_count])
        writer.writerow(["positions_total", position_count])
        writer.writerow(["correct_assignments", correct_assignments])
        writer.writerow(["outdated_templates", len(outdated)])
        writer.writerow(["positions_without_assignment", len(missing_assignment)])
        writer.writerow(["assigned_not_active", len(assigned_not_active)])
        writer.writerow(["legacy_base_codes", len(legacy_base_codes)])
        writer.writerow(["unexpected_question_types", len(unexpected_types)])
        resp = HttpResponse(output.getvalue(), content_type="text/csv; charset=utf-8")
        resp["Content-Disposition"] = 'attachment; filename="system_status.csv"'
        return resp

    return render(
        request,
        "evaluations/report_system.html",
        {
            "now": now,
            "active_count": active_count,
            "position_count": position_count,
            "correct_assignments": correct_assignments,
            "outdated": outdated,
            "q_types": q_types,
            "count_min": count_min,
            "count_max": count_max,
            "count_values": count_values,
            "missing_assignment": missing_assignment,
            "assigned_not_active": assigned_not_active,
            "legacy_base_codes": legacy_base_codes,
            "issues": issues,
        },
    )
